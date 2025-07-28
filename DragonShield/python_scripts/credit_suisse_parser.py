# python_scripts/credit_suisse_parser.py

# MARK: - Version 0.11
# MARK: - History
# - 0.9 -> 0.10: Added CSV support and institution metadata.
# - 0.10 -> 0.11: Return explicit exit codes on errors.

import sys
import re
import openpyxl
import json
import os
import csv
import sqlite3
from datetime import datetime
from dateutil.parser import parse as parse_date
from typing import Dict, Tuple, Any, List, Set, Optional
from openpyxl.cell import Cell, MergedCell # Import Cell types for isinstance checks

# --- Configuration Section (Keep as is) ---
ANLAGEKATEGORIE_TO_GROUP_MAP: Dict[str, str] = {
    "Liquidität & ähnliche": "Cash & Money Market",
    "Festverzinsliche & ähnliche": "Bonds",
    "Aktien & ähnliche": "Equities",
    "AI, Rohstoffe & Immobilien": "Commodities"
}
ASSET_UNTERKATEGORIE_REFINEMENT_MAP: Dict[Tuple[str, str], str] = {
    ("Festverzinsliche & ähnliche", "Obligationenfonds / USD"): "ETFs",
    ("Liquidität & ähnliche", "Geldmarktfonds / CHF"): "Mutual Funds",
    ("Aktien & ähnliche", "Aktienfonds / USA"): "ETFs"
}
COL_ANLAGEKATEGORIE = "Anlagekategorie"
COL_ASSET_UNTERKATEGORIE = "Asset-Unterkategorie"
COL_BESCHREIBUNG = "Beschreibung"
COL_ISIN = "ISIN"
COL_SYMBOL = "Symbol"
COL_VALOR = "Valor"
COL_ANZAHL_NOMINAL = "Anzahl / Nominal"
COL_KOSTEN_KURS = "Einstandskurs"
COL_KOSTEN_KURS_WHRG = "Währung(Einstandskurs)"
COL_AKTUELLER_KURS = "Kurs"
COL_WERT_CHF = "Wert in CHF"
COL_BRANCHE = "Branche"
COL_FAELLIGKEIT = "Fälligkeit"
COL_DEVISENKURS = "Devisenkurs"
COL_DATUM_ZEIT_KURS = "Datum/Zeit des Kurses (Ortszeit der Börse)"

HEADER_ROW_NUMBER = 8
LINE_6_PORTFOLIO_NR_LINE_NUMBER = 6
IDX_WHRG_NOMINAL_FALLBACK = 2 # Fallback if specific header isn't found by name
IDX_WHRG_KURS_FALLBACK = 7    # Fallback

# Default database path can be overridden by the DRAGONSHIELD_DB_PATH environment variable
DB_PATH = os.environ.get(
    "DRAGONSHIELD_DB_PATH",
    "/Users/renekeller/Library/Containers/com.rene.DragonShield/Data/Library/Application Support/DragonShield/dragonshield.sqlite",
)

def _get_actual_cell_value(cell_content: Any) -> Any:
    """Helper to get the .value if it's a Cell object, otherwise return as is."""
    if isinstance(cell_content, (Cell, MergedCell)):
        return cell_content.value
    return cell_content

def parse_statement_date_from_filename(filename: str) -> Optional[str]:
    # (Same as before)
    match_dmY = re.search(r'(\d{1,2})[.-](\d{1,2})[.-](\d{4})', filename)
    if match_dmY:
        day, month, year = match_dmY.groups(); # noinspection PyBroadException
        try: return datetime(int(year), int(month), int(day)).strftime('%Y-%m-%d')
        except Exception: pass
    match_Ymd = re.search(r'(\d{4})[.-]?(\d{2})[.-]?(\d{2})', filename)
    if match_Ymd:
        year, month, day = match_Ymd.groups(); # noinspection PyBroadException
        if len(year) == 4 and len(month) == 2 and len(day) == 2:
             try: return datetime(int(year), int(month), int(day)).strftime('%Y-%m-%d')
             except Exception: pass
    match_mon_dd_yyyy = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s.-]+(\d{1,2})[\s.,-]+(\d{4})', filename, re.IGNORECASE)
    if match_mon_dd_yyyy:
        month_str, day, year = match_mon_dd_yyyy.groups()
        month_map = {name: num for num, name in enumerate(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], 1)}; # noinspection PyBroadException
        month = month_map.get(month_str[:3].capitalize())
        if month:
            try: return datetime(int(year), month, int(day)).strftime('%Y-%m-%d')
            except Exception: pass
    try:
        return parse_date(filename, dayfirst=True).date().isoformat()
    except Exception:
        pass
    return None

def parse_date_from_excel_cell(cell_content: Any, input_format: Optional[str] = None) -> Optional[str]:
    val = _get_actual_cell_value(cell_content)
    if val is None: return None
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        cell_str = val.strip()
        if not cell_str: return None
        if input_format:
            try: return datetime.strptime(cell_str, input_format).strftime('%Y-%m-%d')
            except ValueError: pass
        match_dmY_short_long = re.match(r'(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4})', cell_str)
        if match_dmY_short_long:
            d, m, y_part = match_dmY_short_long.groups(); year = int(y_part)
            if len(y_part) == 2: year += 2000; # noinspection PyBroadException
            try: return datetime(year, int(m), int(d)).strftime('%Y-%m-%d')
            except Exception: pass
        if len(cell_str) >= 10 and cell_str[4] == '-' and cell_str[7] == '-': # YYYY-MM-DD or longer
            try: return datetime.strptime(cell_str[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
            except ValueError: pass
        try:
            return parse_date(cell_str, dayfirst=True).date().isoformat()
        except Exception:
            pass
        # print(f"Warning: Could not parse date string '{cell_str}' to YYYY-MM-DD") # Reduce noise
    return None

def parse_portfolio_nr_from_cell_value(cell_content: Any) -> Optional[str]:
    val = _get_actual_cell_value(cell_content)
    if not isinstance(val, str): return None
    line_content = val.strip()
    match_user_format = re.search(r"S \d{6}-\d{2}", line_content)
    if match_user_format: return match_user_format.group(0)
    match_generic = re.search(r"S\s*\d{6}-\d{2}", line_content)
    if match_generic: return match_generic.group(0)
    return None

def parse_number_from_cell_value(cell_content: Any) -> Optional[float]:
    val = _get_actual_cell_value(cell_content)
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        value_str = val.strip()
        if not value_str: return None
        try:
            cleaned_str = value_str.replace("'", "")
            if "%" in cleaned_str: return float(cleaned_str.replace("%", "")) / 100.0
            return float(cleaned_str)
        except ValueError: # print(f"Warning: Could not parse number from string '{value_str}'") # Reduce noise
            return None
    # print(f"Warning: Unexpected type for number parsing '{type(val)}', value '{val}'") # Reduce noise
    return None

def get_mapped_instrument_group(anlagekategorie: str, asset_unterkategorie: str, unmapped_pairs: Set[Tuple[str, str]]) -> str:
    # (Same as before)
    norm_anlage = anlagekategorie.strip() if anlagekategorie else ""
    norm_unter = asset_unterkategorie.strip() if asset_unterkategorie else ""
    refined_group = ASSET_UNTERKATEGORIE_REFINEMENT_MAP.get((norm_anlage, norm_unter))
    if refined_group: return refined_group
    primary_group = ANLAGEKATEGORIE_TO_GROUP_MAP.get(norm_anlage)
    if primary_group: return primary_group
    if norm_anlage or norm_unter: unmapped_pairs.add((norm_anlage, norm_unter))
    return f"UNMAPPED_CATEGORY"

# --- Instrument lookup helpers ---
def _sanitize(text: str) -> str:
    return "".join(c for c in text if c.isalnum()).upper()

def find_instrument_id_by_valor(conn: sqlite3.Connection, valor: str) -> Optional[int]:
    sanitized_search = _sanitize(valor)
    cur = conn.execute("SELECT instrument_id, valor_nr FROM Instruments WHERE valor_nr IS NOT NULL")
    for inst_id, db_valor in cur.fetchall():
        if _sanitize(db_valor) == sanitized_search:
            return inst_id
    return None

def find_instrument_id_by_isin(conn: sqlite3.Connection, isin: str) -> Optional[int]:
    sanitized_search = _sanitize(isin)
    cur = conn.execute("SELECT instrument_id, isin FROM Instruments WHERE isin IS NOT NULL")
    for inst_id, db_isin in cur.fetchall():
        if _sanitize(db_isin) == sanitized_search:
            return inst_id
    return None

def lookup_instrument_id(conn: Optional[sqlite3.Connection], name: str, valor: str, isin: str) -> Tuple[Optional[int], str]:
    if conn is None:
        return None, ""
    if valor:
        val_id = find_instrument_id_by_valor(conn, valor)
        if val_id is not None:
            return val_id, "Valor"
    if isin:
        isin_id = find_instrument_id_by_isin(conn, isin)
        if isin_id is not None:
            return isin_id, "ISIN"
    return None, ""

# Exit codes for command-line usage
EXIT_SUCCESS = 0
EXIT_FILE_NOT_FOUND = 1
EXIT_DEPENDENCY_ERROR = 2
EXIT_GENERAL_ERROR = 3

def process_file(filepath: str, sheet_name_or_index: Optional[Any] = None) -> int:
    # (Initialization of parsed_data and stats variables remains the same)
    parsed_data = {
        "main_custody_account_nr": None,
        "institution_name": "Credit-Suisse",
        "parsed_statement_date": parse_statement_date_from_filename(filepath.split('/')[-1]),  # Pass only filename
        "summary": {
            "processed_file": filepath,
            "total_data_rows_attempted": 0,
            "data_rows_successfully_parsed": 0,
            "skipped_footer_empty_rows": 0,
            "cash_account_records": 0,
            "security_holding_records": 0,
            "instruments_with_isin": 0,
            "instruments_with_cost_price": 0,
            "unmatched_categories": [],
            "unmatched_instruments": 0,
        },
        "records": [],
        "logs": [],
    }
    main_custody_account_nr_internal: Optional[str] = None
    unmapped_category_pairs_internal: Set[Tuple[str,str]] = set()
    unmatched_instruments_internal = 0

    conn: Optional[sqlite3.Connection] = None
    if os.path.exists(DB_PATH):
        try:
            conn = sqlite3.connect(DB_PATH)
        except Exception as e:
            parsed_data["logs"].append(f"Failed to open database at {DB_PATH}: {e}")
    else:
        parsed_data["logs"].append(f"Database not found at {DB_PATH}; instrument lookup skipped")

    exit_code = EXIT_SUCCESS
    try:
        rows: List[List[Any]]
        if filepath.lower().endswith('.csv'):
            with open(filepath, newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = [list(r) for r in reader]
            sheet = None
            max_column = max(len(r) for r in rows) if rows else 0
            def cell(row: int, col: int):
                return rows[row-1][col-1] if 0 <= row-1 < len(rows) and 0 <= col-1 < len(rows[row-1]) else None
        else:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook[sheet_name_or_index] if sheet_name_or_index is not None and isinstance(sheet_name_or_index, str) else \
                    workbook.worksheets[sheet_name_or_index] if sheet_name_or_index is not None and isinstance(sheet_name_or_index, int) else \
                    workbook.active
            max_column = sheet.max_column
            def cell(row: int, col: int):
                return sheet.cell(row=row, column=col).value
        
        # (Portfolio Nr extraction logic remains the same)
        for col_idx in range(1, min(max_column + 1, 6)):
            cell_val_line6 = cell(LINE_6_PORTFOLIO_NR_LINE_NUMBER, col_idx)
            parsed_nr = parse_portfolio_nr_from_cell_value(cell_val_line6)
            if parsed_nr:
                main_custody_account_nr_internal = parsed_nr
                parsed_data["main_custody_account_nr"] = main_custody_account_nr_internal
                break
        if not main_custody_account_nr_internal:
            print(f"Warning: Could not parse Account Nr from Line {LINE_6_PORTFOLIO_NR_LINE_NUMBER}.\n")


        if sheet is None:
            header_cells = [cell(HEADER_ROW_NUMBER, i) for i in range(1, max_column + 1)]
            headers = [str(h).strip() if h is not None else "" for h in header_cells]
        else:
            header_cells = [c for c in sheet[HEADER_ROW_NUMBER]]
            headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
        header_map = {name: idx for idx, name in enumerate(headers)}
        
        whrg_indices = [i for i, h_name in enumerate(headers) if h_name == "Whrg."]
        idx_whrg_nominal = whrg_indices[0] if len(whrg_indices) > 0 else IDX_WHRG_NOMINAL_FALLBACK
        idx_whrg_kurs = whrg_indices[1] if len(whrg_indices) > 1 else IDX_WHRG_KURS_FALLBACK
        if idx_whrg_nominal == -1: print("Warning: Could not find first 'Whrg.' column header for nominal currency.")
        if idx_whrg_kurs == -1: print("Warning: Could not find second 'Whrg.' column header for price currency.")


        if sheet is None:
            row_iter = enumerate(rows[HEADER_ROW_NUMBER:], start=HEADER_ROW_NUMBER + 1)
        else:
            row_iter = ((idx, [_get_actual_cell_value(c) for c in row])
                        for idx, row in enumerate(sheet.iter_rows(min_row=HEADER_ROW_NUMBER + 1), start=HEADER_ROW_NUMBER + 1))

        for row_idx_iter, row_cells_tuple in row_iter:

            parsed_data["summary"]["total_data_rows_attempted"] += 1
            
            def get_str_val_from_tuple(col_name, default_val=""):
                col_idx = header_map.get(col_name)
                val = row_cells_tuple[col_idx] if col_idx is not None and col_idx < len(row_cells_tuple) else None
                return str(val).strip() if val is not None else default_val

            def get_raw_val_from_tuple(col_name): # Already returns primitive from values_only=True usually, but now from cell.value
                col_idx = header_map.get(col_name)
                if col_idx is not None and col_idx < len(row_cells_tuple): return row_cells_tuple[col_idx]
                return None

            anlagekategorie_str = get_str_val_from_tuple(COL_ANLAGEKATEGORIE)
            beschreibung_str = get_str_val_from_tuple(COL_BESCHREIBUNG)

            if (not anlagekategorie_str and not beschreibung_str and 
                all(c is None or str(c).strip() == "" for c in row_cells_tuple[:min(5, len(row_cells_tuple))])):
                parsed_data["summary"]["skipped_footer_empty_rows"] +=1
                continue
            if len(anlagekategorie_str) > 100 or "real-time daten" in anlagekategorie_str.lower():
                parsed_data["summary"]["skipped_footer_empty_rows"] +=1
                continue

            parsed_data["summary"]["data_rows_successfully_parsed"] += 1
            record_data: Dict[str, Any] = {}
            record_data["institution_name"] = "Credit-Suisse"
            record_data["main_custody_account_nr_from_file"] = main_custody_account_nr_internal
            asset_unterkategorie_str = get_str_val_from_tuple(COL_ASSET_UNTERKATEGORIE)
            mapped_group = get_mapped_instrument_group(anlagekategorie_str, asset_unterkategorie_str, unmapped_category_pairs_internal)
            
            record_data["original_anlagekategorie"] = anlagekategorie_str
            record_data["original_asset_unterkategorie"] = asset_unterkategorie_str
            record_data["mapped_instrument_group_name"] = mapped_group
            record_data["instrument_name_from_file"] = beschreibung_str

            if asset_unterkategorie_str == "Konten":
                # (Cash account processing logic largely the same, ensuring values from tuple are used)
                parsed_data["summary"]["cash_account_records"] += 1
                record_data["record_type"] = "cash_account"
                record_data["cash_account_number_from_file"] = get_str_val_from_tuple(COL_VALOR)
                record_data["currency"] = str(row_cells_tuple[idx_whrg_nominal]).strip() if idx_whrg_nominal != -1 and idx_whrg_nominal < len(row_cells_tuple) and row_cells_tuple[idx_whrg_nominal] is not None else None
                record_data["balance"] = parse_number_from_cell_value(get_raw_val_from_tuple(COL_ANZAHL_NOMINAL))
                record_data["value_in_chf"] = parse_number_from_cell_value(get_raw_val_from_tuple(COL_WERT_CHF))
                record_data["fx_rate_to_chf"] = parse_number_from_cell_value(get_raw_val_from_tuple(COL_DEVISENKURS))
                record_data["asset_class_code"] = "LIQ"
                record_data["asset_sub_class_code"] = "CASH"
            else: 
                # (Security holding processing logic largely the same)
                parsed_data["summary"]["security_holding_records"] += 1
                record_data["record_type"] = "security_holding"
                record_data["main_custody_account_nr_from_file"] = main_custody_account_nr_internal
                record_data["isin"] = get_str_val_from_tuple(COL_ISIN)
                if record_data["isin"]:
                    parsed_data["summary"]["instruments_with_isin"] += 1
                record_data["symbol"] = get_str_val_from_tuple(COL_SYMBOL)
                record_data["valor_nr"] = get_str_val_from_tuple(COL_VALOR)
                record_data["quantity_nominal"] = parse_number_from_cell_value(get_raw_val_from_tuple(COL_ANZAHL_NOMINAL))
                cost_price_val = get_raw_val_from_tuple(COL_KOSTEN_KURS)
                record_data["cost_price"] = parse_number_from_cell_value(cost_price_val)
                if record_data["cost_price"] is not None: parsed_data["summary"]["instruments_with_cost_price"] +=1 # Corrected counter
                record_data["cost_price_currency"] = get_str_val_from_tuple(COL_KOSTEN_KURS_WHRG)
                current_price_val = get_raw_val_from_tuple(COL_AKTUELLER_KURS)
                record_data["current_price"] = parse_number_from_cell_value(current_price_val)
                record_data["current_price_currency"] = str(row_cells_tuple[idx_whrg_kurs]).strip() if idx_whrg_kurs != -1 and idx_whrg_kurs < len(row_cells_tuple) and row_cells_tuple[idx_whrg_kurs] is not None else None
                record_data["value_in_chf"] = parse_number_from_cell_value(get_raw_val_from_tuple(COL_WERT_CHF))
                record_data["sector"] = get_str_val_from_tuple(COL_BRANCHE)
                maturity_date_val = get_raw_val_from_tuple(COL_FAELLIGKEIT)
                record_data["maturity_date"] = parse_date_from_excel_cell(maturity_date_val, input_format='%d.%m.%y')
                price_date_val = get_raw_val_from_tuple(COL_DATUM_ZEIT_KURS)
                record_data["price_date"] = parse_date_from_excel_cell(price_date_val)

                valor_col = row_cells_tuple[5] if len(row_cells_tuple) > 5 else None
                isin_col = row_cells_tuple[22] if len(row_cells_tuple) > 22 else None
                valor_str = str(valor_col).strip() if valor_col is not None else ""
                isin_str = str(isin_col).strip() if isin_col is not None else ""
                instr_id, method = lookup_instrument_id(conn, beschreibung_str, valor_str, isin_str)

                if instr_id is not None:
                    record_data["instrument_id"] = instr_id
                    log_msg = (
                        f"Matched instrument {beschreibung_str} (ID: {instr_id}) via {method} {valor_str if method == 'Valor' else isin_str} "
                        f"| Valor: {valor_str or 'N/A'}, ISIN: {isin_str or 'N/A'}"
                    )
                else:
                    unmatched_instruments_internal += 1
                    log_msg = (
                        f"Unmatched instrument description: {beschreibung_str} "
                        f"| Valor: {valor_str or 'N/A'}, ISIN: {isin_str or 'N/A'}"
                    )
                parsed_data["logs"].append(log_msg)
            
            parsed_data["records"].append(record_data)
        
        parsed_data["summary"]["unmapped_categories"] = sorted(list(unmapped_category_pairs_internal))
        parsed_data["summary"]["unmatched_instruments"] = unmatched_instruments_internal

        if conn is not None:
            conn.close()

    # (Exception handling and JSON printing remain the same)
    except FileNotFoundError:
        parsed_data["summary"]["error"] = f"File not found at {filepath}"
        exit_code = EXIT_FILE_NOT_FOUND
    except ImportError:
        parsed_data["summary"]["error"] = "The 'openpyxl' library is required. Please install it (e.g., pip install openpyxl)."
        exit_code = EXIT_DEPENDENCY_ERROR
    except Exception as e:
        import traceback
        parsed_data["summary"]["error"] = f"An error occurred: {str(e)}"
        parsed_data["summary"]["traceback"] = traceback.format_exc()
        exit_code = EXIT_GENERAL_ERROR
    
    print(json.dumps(parsed_data, indent=2, ensure_ascii=False))
    return exit_code

if __name__ == "__main__":
    if len(sys.argv) > 1:
        filepath_arg = sys.argv[1]
        code = process_file(filepath_arg)
        sys.exit(code)
    else:
        print(json.dumps({"error": "Please provide the XLSX file path as an argument.", "usage": "python credit_suisse_parser.py <path_to_your_Credit-Suisse_file.xlsx>"}))
        sys.exit(1)

