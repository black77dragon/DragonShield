# python_scripts/zkb_parser.py
import sys
import re
import openpyxl
from typing import Dict, Tuple, Any, List, Set, Optional

# --- Configuration Section ---
ANLAGEKATEGORIE_TO_GROUP_MAP: Dict[str, str] = {
    "Liquidität & ähnliche": "Cash & Money Market",
    "Festverzinsliche & ähnliche": "Bonds",
    "Aktien & ähnliche": "Equities",
    "AI, Rohstoffe & Immobilien": "Commodities" # Added based on user feedback
}

ASSET_UNTERKATEGORIE_REFINEMENT_MAP: Dict[Tuple[str, str], str] = {
    ("Festverzinsliche & ähnliche", "Obligationenfonds / USD"): "ETFs",
    ("Liquidität & ähnliche", "Geldmarktfonds / CHF"): "Mutual Funds",
    ("Aktien & ähnliche", "Aktienfonds / USA"): "ETFs" # Added based on user feedback for Row 29
}

# Key column names (ensure these exactly match your CSV/Excel headers)
COL_ANLAGEKATEGORIE = "Anlagekategorie"
COL_ASSET_UNTERKATEGORIE = "Asset-Unterkategorie"
COL_BESCHREIBUNG = "Beschreibung"
COL_ISIN = "ISIN"
COL_SYMBOL = "Symbol"
COL_VALOR = "Valor"
COL_ANZAHL_NOMINAL = "Anzahl / Nominal"
COL_KOSTEN_KURS = "Einstandskurs"
COL_KOSTEN_KURS_WHRG = "Währung(Einstandskurs)"
COL_AKTUELLER_KURS = "Kurs"
COL_WERT_CHF = "Wert in CHF"
COL_BRANCHE = "Branche"
COL_FAELLIGKEIT = "Fälligkeit"
COL_DEVISENKURS = "Devisenkurs"
# For potentially duplicated headers like "Whrg.", we'll rely on indices
# If header names are unique in the source, direct name access is fine.

HEADER_ROW_NUMBER = 8 # Line number in Excel where headers start (1-based)
LINE_6_PORTFOLIO_NR_LINE_NUMBER = 6 # Line number in Excel for Portfolio Nr. (1-based)

# Expected indices for 'Whrg.' columns if headers are ambiguous (0-based)
# Whrg. (for Anzahl/Nominal) is header[2]
# Whrg. (for Kurs) is header[7]
IDX_WHRG_NOMINAL_FALLBACK = 2
IDX_WHRG_KURS_FALLBACK = 7


def parse_portfolio_nr_from_cell_value(cell_value: Any) -> Optional[str]:
    if not isinstance(cell_value, str):
        return None
    line_content = cell_value.strip()
    match_user_format = re.search(r"S \d{6}-\d{2}", line_content)
    if match_user_format:
         return match_user_format.group(0)
    match_generic = re.search(r"S\s*\d{6}-\d{2}", line_content)
    if match_generic:
        return match_generic.group(0)
    return None

def parse_number_from_cell_value(cell_value: Any) -> Optional[float]:
    if cell_value is None: return None
    if isinstance(cell_value, (int, float)): return float(cell_value)
    if isinstance(cell_value, str):
        value_str = cell_value.strip()
        if not value_str: return None
        try:
            cleaned_str = value_str.replace("'", "") # Remove ' as thousands separator
            if "%" in cleaned_str:
                return float(cleaned_str.replace("%", "")) / 100.0
            return float(cleaned_str)
        except ValueError:
            print(f"Warning: Could not parse number from string '{value_str}'")
            return None
    print(f"Warning: Unexpected type for number parsing '{type(cell_value)}', value '{cell_value}'")
    return None

def get_mapped_instrument_group(anlagekategorie: str, asset_unterkategorie: str, unmapped_pairs: Set[Tuple[str, str]]) -> str:
    norm_anlage = anlagekategorie.strip() if anlagekategorie else ""
    norm_unter = asset_unterkategorie.strip() if asset_unterkategorie else ""

    refined_group = ASSET_UNTERKATEGORIE_REFINEMENT_MAP.get((norm_anlage, norm_unter))
    if refined_group: return refined_group
    
    primary_group = ANLAGEKATEGORIE_TO_GROUP_MAP.get(norm_anlage)
    if primary_group: return primary_group
    
    if norm_anlage or norm_unter: # Add if there was some category info
        unmapped_pairs.add((norm_anlage, norm_unter))
    return f"UNMAPPED_CATEGORY"

def process_file(filepath: str, sheet_name_or_index: Optional[Any] = None):
    print(f"Processing Excel Statement: {filepath}\n")
    main_custody_account_nr: Optional[str] = None
    
    total_data_rows_processed = 0
    cash_records_count = 0
    security_records_count = 0
    unmapped_category_pairs: Set[Tuple[str,str]] = set()
    instruments_with_isin = 0
    instruments_with_cost_price = 0
    skipped_footer_rows = 0

    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook[sheet_name_or_index] if sheet_name_or_index is not None and isinstance(sheet_name_or_index, str) else \
                workbook.worksheets[sheet_name_or_index] if sheet_name_or_index is not None and isinstance(sheet_name_or_index, int) else \
                workbook.active
        
        print(f"Reading from sheet: '{sheet.title}'")

        # Read Line 6 for Portfolio Nr.
        for col_idx in range(1, min(sheet.max_column + 1, 6)):
            cell_value_line6 = sheet.cell(row=LINE_6_PORTFOLIO_NR_LINE_NUMBER, column=col_idx).value
            if cell_value_line6 and isinstance(cell_value_line6, str):
                parsed_nr = parse_portfolio_nr_from_cell_value(cell_value_line6)
                if parsed_nr:
                    main_custody_account_nr = parsed_nr
                    print(f"--- Extracted Main Custody Account Nr (Line {LINE_6_PORTFOLIO_NR_LINE_NUMBER}, Col {col_idx}): {main_custody_account_nr} ---\n")
                    break 
        if not main_custody_account_nr:
            print(f"Warning: Could not parse Custody Account Nr from Line {LINE_6_PORTFOLIO_NR_LINE_NUMBER}.\n")

        # Read Headers from HEADER_ROW_NUMBER
        header_cells = [cell for cell in sheet[HEADER_ROW_NUMBER]] # Get cell objects
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        
        # Create a map of header names to their column index (0-based)
        header_map = {name: idx for idx, name in enumerate(headers)}
        print(f"Detected Headers on Line {HEADER_ROW_NUMBER}: {headers}\n")

        # Identify indices of 'Whrg.' columns for robust access
        # This assumes the text "Whrg." is constant.
        whrg_indices = [i for i, h_name in enumerate(headers) if h_name == "Whrg."]
        idx_whrg_nominal = whrg_indices[0] if len(whrg_indices) > 0 else -1
        idx_whrg_kurs = whrg_indices[1] if len(whrg_indices) > 1 else -1
        
        if idx_whrg_nominal == -1: print("Warning: Could not find first 'Whrg.' column header for nominal currency.")
        if idx_whrg_kurs == -1: print("Warning: Could not find second 'Whrg.' column header for price currency.")


        # Iterate through data rows starting from HEADER_ROW_NUMBER + 1
        for row_idx, row_cells_tuple in enumerate(sheet.iter_rows(min_row=HEADER_ROW_NUMBER + 1, values_only=True), start=HEADER_ROW_NUMBER + 1):
            
            # Helper to get cell value by mapped header name, or by index from tuple
            def get_cell_value_by_name(col_name, default_val=""):
                col_idx = header_map.get(col_name)
                if col_idx is not None and col_idx < len(row_cells_tuple):
                    return row_cells_tuple[col_idx]
                return default_val

            anlagekategorie_val = get_cell_value_by_name(COL_ANLAGEKATEGORIE)
            beschreibung_val = get_cell_value_by_name(COL_BESCHREIBUNG)

            # Skip potential footer/disclaimer rows
            anlagekategorie_str = str(anlagekategorie_val).strip() if anlagekategorie_val is not None else ""
            beschreibung_str = str(beschreibung_val).strip() if beschreibung_val is not None else ""

            if not anlagekategorie_str and not beschreibung_str and len(row_cells_tuple) > 0 and all(c is None or str(c).strip() == "" for c in row_cells_tuple[:5]): # Example: if first 5 cells are effectively empty
                skipped_footer_rows +=1
                # print(f"Skipping likely empty/footer row {row_idx}")
                continue
            if len(anlagekategorie_str) > 100 or "real-time daten" in anlagekategorie_str.lower(): # Skip long disclaimer in Anlagekategorie
                skipped_footer_rows +=1
                # print(f"Skipping likely disclaimer row {row_idx} based on Anlagekategorie content.")
                continue


            total_data_rows_processed += 1
            print(f"--- Row {row_idx} Data ---")
            
            asset_unterkategorie_val = get_cell_value_by_name(COL_ASSET_UNTERKATEGORIE)
            asset_unterkategorie_str = str(asset_unterkategorie_val).strip() if asset_unterkategorie_val is not None else ""
            
            mapped_group = get_mapped_instrument_group(anlagekategorie_str, asset_unterkategorie_str, unmapped_category_pairs)
            
            print(f"  Original Anlagekategorie: '{anlagekategorie_str}'")
            print(f"  Original Asset-Unterkategorie: '{asset_unterkategorie_str}'")
            print(f"  Mapped Instrument Group: '{mapped_group}'")
            print(f"  Instrument Name (Beschreibung): '{beschreibung_str}'")

            if asset_unterkategorie_str == "Konten":
                cash_records_count += 1
                print(f"  Record Type: Cash Account")
                cash_acc_nr = str(get_cell_value_by_name(COL_VALOR)).strip()
                cash_currency = str(row_cells_tuple[idx_whrg_nominal]).strip() if idx_whrg_nominal != -1 and idx_whrg_nominal < len(row_cells_tuple) else "N/A"
                
                balance = parse_number_from_cell_value(get_cell_value_by_name(COL_ANZAHL_NOMINAL))
                wert_chf = parse_number_from_cell_value(get_cell_value_by_name(COL_WERT_CHF))
                devisenkurs = parse_number_from_cell_value(get_cell_value_by_name(COL_DEVISENKURS))

                print(f"  Cash Account Number (Valor): '{cash_acc_nr}'")
                print(f"  Balance: {balance if balance is not None else 'N/A'} {cash_currency}")
                print(f"  Value in CHF: {wert_chf if wert_chf is not None else 'N/A'}")
                if devisenkurs is not None: print(f"  FX Rate (Devisenkurs): {devisenkurs}")
            
            else: 
                security_records_count += 1
                print(f"  Record Type: Security/Fund Holding")
                if main_custody_account_nr: print(f"  Belongs to Custody Account: {main_custody_account_nr}")
                else: print(f"  Warning: Main Custody Account Nr not available for this security.")

                isin = str(get_cell_value_by_name(COL_ISIN)).strip()
                symbol = str(get_cell_value_by_name(COL_SYMBOL)).strip()
                valor_instrument = str(get_cell_value_by_name(COL_VALOR)).strip()
                
                if isin: instruments_with_isin += 1
                if isin: print(f"  ISIN: '{isin}'")
                if symbol: print(f"  Symbol: '{symbol}'")
                if valor_instrument: print(f"  Valor (Instrument): '{valor_instrument}'")

                quantity = parse_number_from_cell_value(get_cell_value_by_name(COL_ANZAHL_NOMINAL))
                
                cost_price_val = get_cell_value_by_name(COL_KOSTEN_KURS)
                cost_price = parse_number_from_cell_value(cost_price_val)
                if cost_price is not None : instruments_with_cost_price +=1
                cost_price_currency = str(get_cell_value_by_name(COL_KOSTEN_KURS_WHRG)).strip()
                
                current_price_val = get_cell_value_by_name(COL_AKTUELLER_KURS)
                current_price = parse_number_from_cell_value(current_price_val) # Handles '%'
                price_currency = str(row_cells_tuple[idx_whrg_kurs]).strip() if idx_whrg_kurs != -1 and idx_whrg_kurs < len(row_cells_tuple) else "N/A"
                
                wert_chf = parse_number_from_cell_value(get_cell_value_by_name(COL_WERT_CHF))
                sector = str(get_cell_value_by_name(COL_BRANCHE)).strip()
                maturity_val = get_cell_value_by_name(COL_FAELLIGKEIT)
                maturity = str(maturity_val).strip() if maturity_val is not None else ""


                print(f"  Quantity/Nominal: {quantity if quantity is not None else 'N/A'}")
                print(f"  Cost Price (Einstandskurs): {cost_price if cost_price is not None else 'N/A'} {cost_price_currency}")
                print(f"  Current Price (Kurs): {current_price if current_price is not None else 'N/A'} {price_currency}")
                print(f"  Value in CHF: {wert_chf if wert_chf is not None else 'N/A'}")
                if sector: print(f"  Sector (Branche): '{sector}'")
                if maturity: print(f"  Maturity (Fälligkeit): '{maturity}'")
            
            print("-" * 30 + "\n")

    except FileNotFoundError:
        print(f"Error: File not found at {filepath}")
        return 
    except ImportError:
        print("Error: The 'openpyxl' library is required to read .xlsx files. Please install it (e.g., pip install openpyxl).")
        return
    except Exception as e:
        print(f"An error occurred during file processing: {e}")
        import traceback
        traceback.print_exc()
        return

    print("\n--- IMPORT SUMMARY ---")
    print(f"Processed File: {filepath}")
    if main_custody_account_nr: print(f"Main Custody Account Nr. from File: {main_custody_account_nr}")
    print(f"Total Data Rows Attempted (after header): {total_data_rows_processed + skipped_footer_rows}")
    print(f"Data Rows Successfully Parsed: {total_data_rows_processed}")
    print(f"Skipped Footer/Empty Rows: {skipped_footer_rows}")
    print(f"Cash Account Records Identified: {cash_records_count}")
    print(f"Security/Fund Holding Records Identified: {security_records_count}")
    print(f"Instruments with ISIN found: {instruments_with_isin}")
    print(f"Instruments with Cost Price (Einstandskurs) found: {instruments_with_cost_price}")
    
    if unmapped_category_pairs:
        print(f"\nEncountered {len(unmapped_category_pairs)} Unique Unmapped Category Pair(s):")
        for anlage, unter in sorted(list(unmapped_category_pairs)):
            print(f"  - Anlagekategorie: '{anlage}', Asset-Unterkategorie: '{unter}'")
        print("These will need to be mapped or new Instrument Groups created in the application.")
    else:
        print("\nAll encountered categories were successfully mapped to Instrument Groups.")
    print("--- END OF SUMMARY ---")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        filepath_arg = sys.argv[1]
        process_file(filepath_arg)
    else:
        print("Please provide the XLSX file path as an argument.")
        print("Usage: python zkb_parser.py <path_to_your_ZKB_file.xlsx>")