# python_scripts/zkb_parser.py
import sys
import re
import openpyxl # For reading .xlsx files
from typing import Dict, Tuple, Any, List, Set, Optional

# --- Configuration Section (Keep as is from previous version) ---
ANLAGEKATEGORIE_TO_GROUP_MAP: Dict[str, str] = {
    "Liquidität & ähnliche": "Cash & Money Market",
    "Festverzinsliche & ähnliche": "Bonds",
    "Aktien & ähnliche": "Equities"
}

ASSET_UNTERKATEGORIE_REFINEMENT_MAP: Dict[Tuple[str, str], str] = {
    ("Festverzinsliche & ähnliche", "Obligationenfonds / USD"): "ETFs",
    ("Liquidität & ähnliche", "Geldmarktfonds / CHF"): "Mutual Funds"
}

# Column names (these will be matched against the header row read from Excel)
COL_ANLAGEKATEGORIE = "Anlagekategorie"
COL_ASSET_UNTERKATEGORIE = "Asset-Unterkategorie"
COL_BESCHREIBUNG = "Beschreibung"
COL_ISIN = "ISIN"
COL_SYMBOL = "Symbol"
COL_VALOR = "Valor"
COL_ANZAHL_NOMINAL = "Anzahl / Nominal"
COL_KOSTEN_KURS = "Einstandskurs"
COL_KOSTEN_KURS_WHRG = "Währung(Einstandskurs)"
COL_AKTUELLER_KURS = "Kurs"
COL_WERT_CHF = "Wert in CHF"
COL_BRANCHE = "Branche"
COL_FAELLIGKEIT = "Fälligkeit"
COL_DEVISENKURS = "Devisenkurs"
# We still need to handle the two "Whrg." columns.
# When reading headers from Excel, we'll get them as a list.
# We will identify their indices to reliably access them.
HEADER_WHRG_NOMINAL_TEXT = "Whrg." # Text of the header
HEADER_WHRG_KURS_TEXT = "Whrg."    # Text of the header

HEADER_ROW_NUMBER = 8 # Line number in Excel where headers start (1-based)
LINE_6_PORTFOLIO_NR_LINE_NUMBER = 6 # Line number in Excel for Portfolio Nr. (1-based)

def parse_portfolio_nr_from_cell_value(cell_value: Any) -> Optional[str]:
    if not isinstance(cell_value, str):
        return None
    line_content = cell_value.strip()
    # User confirmed format "Portfolio-Nr. S 398424-05"
    match_user_format = re.search(r"S \d{6}-\d{2}", line_content)
    if match_user_format:
         return match_user_format.group(0) # e.g., "S 398424-05"
    # Fallback for a more generic pattern if needed
    match_generic = re.search(r"S\s*\d{6}-\d{2}", line_content)
    if match_generic: # e.g. S123456-78 or S 123456-78
        return match_generic.group(0) 
    return None

def parse_number_from_cell_value(cell_value: Any) -> Optional[float]:
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float)):
        return float(cell_value)
    if isinstance(cell_value, str):
        value_str = cell_value.strip()
        if not value_str:
            return None
        try:
            cleaned_str = value_str.replace("'", "") # Remove thousand separators
            # Add more cleaning if needed (e.g., for currency symbols, percentage signs if price is "99.50%")
            if "%" in cleaned_str:
                return float(cleaned_str.replace("%", "")) / 100.0
            return float(cleaned_str)
        except ValueError:
            print(f"Warning: Could not parse number from string '{value_str}'")
            return None
    print(f"Warning: Unexpected type for number parsing '{type(cell_value)}', value '{cell_value}'")
    return None

def get_mapped_instrument_group(anlagekategorie: str, asset_unterkategorie: str, unmapped_pairs: Set[Tuple[str, str]]) -> str:
    # (This function remains the same as before)
    norm_anlage = anlagekategorie.strip() if anlagekategorie else ""
    norm_unter = asset_unterkategorie.strip() if asset_unterkategorie else ""

    refined_group = ASSET_UNTERKATEGORIE_REFINEMENT_MAP.get((norm_anlage, norm_unter))
    if refined_group:
        return refined_group
    
    primary_group = ANLAGEKATEGORIE_TO_GROUP_MAP.get(norm_anlage)
    if primary_group:
        return primary_group
    
    if norm_anlage or norm_unter: # Only add if there was some category info
        unmapped_pairs.add((norm_anlage, norm_unter))
    return f"UNMAPPED_CATEGORY"

def process_file(filepath: str, sheet_name_or_index: Optional[Any] = None):
    print(f"Processing Excel Statement: {filepath}\n")
    main_custody_account_nr: Optional[str] = None
    
    # Statistics (same as before)
    total_data_rows_processed = 0
    cash_records_count = 0
    security_records_count = 0
    unmapped_category_pairs: Set[Tuple[str,str]] = set()
    instruments_with_isin = 0
    instruments_with_cost_price = 0

    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True) # data_only=True to get values, not formulas
        
        if sheet_name_or_index is not None:
            if isinstance(sheet_name_or_index, str):
                sheet = workbook[sheet_name_or_index]
            elif isinstance(sheet_name_or_index, int):
                sheet = workbook.worksheets[sheet_name_or_index]
            else:
                sheet = workbook.active # Default to active sheet
        else:
            sheet = workbook.active # Default to active sheet
        
        print(f"Reading from sheet: '{sheet.title}'")

        # Read Line 6 for Portfolio Nr.
        # Assuming the relevant text is in the first few cells of that row.
        # Iterate through cells in row 6 to find the pattern.
        for col_idx in range(1, min(sheet.max_column + 1, 6)): # Check first 5 columns
            cell_value_line6 = sheet.cell(row=LINE_6_PORTFOLIO_NR_LINE_NUMBER, column=col_idx).value
            if cell_value_line6 and isinstance(cell_value_line6, str):
                parsed_nr = parse_portfolio_nr_from_cell_value(cell_value_line6)
                if parsed_nr:
                    main_custody_account_nr = parsed_nr
                    print(f"--- Extracted Main Custody Account Nr (Line {LINE_6_PORTFOLIO_NR_LINE_NUMBER}, Col {col_idx}): {main_custody_account_nr} ---\n")
                    break 
        if not main_custody_account_nr:
            print(f"Warning: Could not parse Custody Account Nr from Line {LINE_6_PORTFOLIO_NR_LINE_NUMBER}.\n")

        # Read Headers from HEADER_ROW_NUMBER
        headers_row_values = [cell.value for cell in sheet[HEADER_ROW_NUMBER]]
        headers = [str(h).strip() if h is not None else "" for h in headers_row_values]
        
        # Identify indices of potentially duplicated 'Whrg.' columns
        whrg_indices = [i for i, h_name in enumerate(headers) if h_name == HEADER_WHRG_NOMINAL_TEXT]
        idx_whrg_nominal = whrg_indices[0] if len(whrg_indices) > 0 else -1 # Typically the first 'Whrg.'
        idx_whrg_kurs = whrg_indices[1] if len(whrg_indices) > 1 else -1    # Typically the second 'Whrg.'

        print(f"Detected Headers on Line {HEADER_ROW_NUMBER}: {headers}\n")

        # Create a map of header names to their column index for easy lookup
        header_map = {name: idx for idx, name in enumerate(headers)}

        # Iterate through data rows starting from HEADER_ROW_NUMBER + 1
        for row_idx, row_cells_tuple in enumerate(sheet.iter_rows(min_row=HEADER_ROW_NUMBER + 1, values_only=True), start=HEADER_ROW_NUMBER + 1):
            # Convert tuple of cell values to a dictionary using headers for easier access
            # This assumes headers are unique enough after stripping.
            # If headers are not unique, this dict might lose data.
            # It's safer to use row_cells_tuple directly with indices from header_map.
            
            total_data_rows_processed += 1
            print(f"--- Row {row_idx} Data ---")

            def get_value(col_name, default=""):
                idx = header_map.get(col_name)
                if idx is not None and idx < len(row_cells_tuple):
                    return str(row_cells_tuple[idx]).strip() if row_cells_tuple[idx] is not None else default
                return default

            anlagekategorie = get_value(COL_ANLAGEKATEGORIE)
            asset_unterkategorie = get_value(COL_ASSET_UNTERKATEGORIE)
            
            mapped_group = get_mapped_instrument_group(anlagekategorie, asset_unterkategorie, unmapped_category_pairs)
            
            print(f"  Original Anlagekategorie: '{anlagekategorie}'")
            print(f"  Original Asset-Unterkategorie: '{asset_unterkategorie}'")
            print(f"  Mapped Instrument Group: '{mapped_group}'")

            beschreibung = get_value(COL_BESCHREIBUNG, "N/A")
            
            if asset_unterkategorie == "Konten":
                cash_records_count += 1
                print(f"  Record Type: Cash Account")
                cash_acc_nr = get_value(COL_VALOR)
                cash_currency = str(row_cells_tuple[idx_whrg_nominal]).strip() if idx_whrg_nominal != -1 and idx_whrg_nominal < len(row_cells_tuple) else "N/A"
                
                balance = parse_number_from_cell_value(row_cells_tuple[header_map.get(COL_ANZAHL_NOMINAL)] if header_map.get(COL_ANZAHL_NOMINAL) is not None and header_map.get(COL_ANZAHL_NOMINAL) < len(row_cells_tuple) else None)
                wert_chf = parse_number_from_cell_value(row_cells_tuple[header_map.get(COL_WERT_CHF)] if header_map.get(COL_WERT_CHF) is not None and header_map.get(COL_WERT_CHF) < len(row_cells_tuple) else None)
                devisenkurs = parse_number_from_cell_value(row_cells_tuple[header_map.get(COL_DEVISENKURS)] if header_map.get(COL_DEVISENKURS) is not None and header_map.get(COL_DEVISENKURS) < len(row_cells_tuple) else None)

                print(f"  Cash Account Name (Beschreibung): '{beschreibung}'")
                print(f"  Cash Account Number (Valor): '{cash_acc_nr}'")
                print(f"  Balance: {balance if balance is not None else 'N/A'} {cash_currency}")
                print(f"  Value in CHF: {wert_chf if wert_chf is not None else 'N/A'}")
                if devisenkurs is not None: print(f"  FX Rate (Devisenkurs): {devisenkurs}")
            
            else: 
                security_records_count += 1
                # (Rest of the security/fund holding processing logic remains similar,
                #  using get_value or row_cells_tuple[header_map.get(...)])
                # Make sure to use idx_whrg_kurs for price_currency
                print(f"  Record Type: Security/Fund Holding")
                # ... (print details for securities as in the previous CSV script version) ...
                # Example for one field:
                isin = get_value(COL_ISIN)
                if isin: 
                    instruments_with_isin += 1
                    print(f"  ISIN: '{isin}'")
                
                cost_price_val = row_cells_tuple[header_map.get(COL_KOSTEN_KURS)] if header_map.get(COL_KOSTEN_KURS) is not None and header_map.get(COL_KOSTEN_KURS) < len(row_cells_tuple) else None
                cost_price = parse_number_from_cell_value(cost_price_val)
                if cost_price is not None: instruments_with_cost_price +=1
                cost_price_currency = get_value(COL_KOSTEN_KURS_WHRG, "N/A")
                print(f"  Cost Price (Einstandskurs): {cost_price if cost_price is not None else 'N/A'} {cost_price_currency}")
                # ... etc. for other fields ...

            print("-" * 30 + "\n")

    except FileNotFoundError:
        print(f"Error: File not found at {filepath}")
        return 
    except ImportError:
        print("Error: The 'openpyxl' library is required to read .xlsx files. Please install it (e.g., pip install openpyxl).")
        return
    except Exception as e:
        print(f"An error occurred during file processing: {e}")
        import traceback
        traceback.print_exc()
        return

    # --- Summary Report (same as before) ---
    print("\n--- IMPORT SUMMARY ---")
    # ...
    print(f"Processed File: {filepath}")
    if main_custody_account_nr:
        print(f"Main Custody Account Nr. from File: {main_custody_account_nr}")
    print(f"Total Data Rows Read (from sheet): {total_data_rows_processed}")
    print(f"Cash Account Records Identified: {cash_records_count}")
    print(f"Security/Fund Holding Records Identified: {security_records_count}")
    print(f"Instruments with ISIN found: {instruments_with_isin}")
    print(f"Instruments with Cost Price (Einstandskurs) found: {instruments_with_cost_price}")
    
    if unmapped_category_pairs:
        print(f"\nEncountered {len(unmapped_category_pairs)} Unique Unmapped Category Pair(s):")
        for anlage, unter in sorted(list(unmapped_category_pairs)):
            print(f"  - Anlagekategorie: '{anlage}', Asset-Unterkategorie: '{unter}'")
        print("These will need to be mapped or new Instrument Groups created in the application.")
    else:
        print("\nAll encountered categories were successfully mapped to Instrument Groups.")
    print("--- END OF SUMMARY ---")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        filepath_arg = sys.argv[1]
        # Example: process_file(filepath_arg, sheet_name_or_index="YourSheetName")
        # or process_file(filepath_arg, sheet_name_or_index=0) for the first sheet
        process_file(filepath_arg) # Defaults to active sheet
    else:
        print("Please provide the XLSX file path as an argument.")
        print("Usage: python zkb_parser.py <path_to_your_ZKB_file.xlsx>")